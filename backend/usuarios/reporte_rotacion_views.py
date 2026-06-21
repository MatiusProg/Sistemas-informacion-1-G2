"""
============================================================
ARCHIVO: backend/usuarios/reporte_rotacion_views.py
CASO DE USO: CU26 - Generar Reporte de Rotacion de Inventario
CICLO: 4
AUTOR: Adalid
FECHA: 21/06/26
============================================================

DESCRIPCION:
Calcula, para cada insumo, los totales de ingreso, salida y merma
registrados en MOVIMIENTO_INVENTARIO dentro del rango de fechas
indicado, y calcula el indice de rotacion (consumido/ingresado*100).
Incluye el stock actual (suma de cantidades en tabla STOCK por insumo).

No se crea ninguna tabla nueva. Se generan archivos PDF (reportlab)
y Excel (openpyxl) reales para descarga. Solo la exportacion deja
constancia en la bitacora (GENERAR_REPORTE_ROTACION).

Tablas consultadas (todas ya existentes):
- MOVIMIENTO_INVENTARIO  (insumo_id, tipo, cantidad, fecha_mov)
- INSUMO                 (id, nombre, categoria)
- STOCK                  (insumo_id, cantidad)
"""

import io
import logging

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from supabase import create_client
from django.conf import settings
from django.http import HttpResponse

from bitacora.utils import registrar_accion, obtener_ip_cliente

logger = logging.getLogger(__name__)


def _calcular_reporte_rotacion(supabase, fecha_desde=None, fecha_hasta=None, insumo_id=None):
    """
    Funcion compartida por los 3 endpoints (JSON, PDF, Excel).
    Calcula total ingresado, consumido, mermado, rotacion y stock
    actual para cada insumo (o para uno solo si se indica insumo_id).

    Retorna una lista de dicts:
    [
        {
            "insumo_id": int,
            "nombre": str,
            "categoria": str,
            "total_ingresado": float,
            "total_consumido": float,
            "total_mermado": float,
            "rotacion": float,
            "stock_actual": float
        },
        ...
    ]
    ordenada por nombre de insumo.
    """
    # 1) Obtener todos los insumos (o uno especifico)
    insumos_query = supabase.table('insumo').select('id, nombre, categoria')
    if insumo_id is not None:
        insumos_query = insumos_query.eq('id', insumo_id)
    insumos = insumos_query.execute().data or []

    if not insumos:
        return []

    insumo_ids = [i['id'] for i in insumos]
    insumos_map = {i['id']: i for i in insumos}

    # 2) Movimientos en el rango de fechas
    movimientos_query = supabase.table('movimiento_inventario').select(
        'insumo_id, tipo, cantidad'
    ).in_('insumo_id', insumo_ids)

    if fecha_desde:
        movimientos_query = movimientos_query.gte('fecha_mov', fecha_desde)
    if fecha_hasta:
        movimientos_query = movimientos_query.lte('fecha_mov', fecha_hasta)
    if insumo_id is not None:
        movimientos_query = movimientos_query.eq('insumo_id', insumo_id)

    movimientos = movimientos_query.execute().data or []

    # 3) Stock actual: suma de cantidades por insumo_id
    stock_query = supabase.table('stock').select('insumo_id, cantidad').in_('insumo_id', insumo_ids)
    stocks = stock_query.execute().data or []

    stock_por_insumo = {}
    for s in stocks:
        iid = s['insumo_id']
        stock_por_insumo[iid] = stock_por_insumo.get(iid, 0.0) + float(s['cantidad'] or 0)

    # 4) Acumular totales por insumo
    totales = {}
    for iid in insumo_ids:
        totales[iid] = {
            'total_ingresado': 0.0,
            'total_consumido': 0.0,
            'total_mermado': 0.0,
        }

    for m in movimientos:
        iid = m['insumo_id']
        if iid not in totales:
            continue
        cantidad = float(m.get('cantidad') or 0)
        tipo = (m.get('tipo') or '').lower()

        if tipo == 'ingreso':
            totales[iid]['total_ingresado'] += cantidad
        elif tipo == 'salida':
            totales[iid]['total_consumido'] += cantidad
        elif tipo == 'merma':
            totales[iid]['total_mermado'] += cantidad

    # 5) Construir reporte
    reporte = []
    for iid in insumo_ids:
        info = insumos_map[iid]
        t = totales[iid]
        ingresado = t['total_ingresado']
        consumido = t['total_consumido']
        mermado = t['total_mermado']

        if ingresado > 0:
            rotacion = round(consumido / ingresado * 100, 2)
        else:
            rotacion = 0.0

        reporte.append({
            'insumo_id': iid,
            'nombre': info.get('nombre', 'Desconocido'),
            'categoria': info.get('categoria', ''),
            'total_ingresado': round(ingresado, 2),
            'total_consumido': round(consumido, 2),
            'total_mermado': round(mermado, 2),
            'rotacion': rotacion,
            'stock_actual': round(stock_por_insumo.get(iid, 0.0), 2),
        })

    reporte.sort(key=lambda x: x['nombre'])
    return reporte


class ReporteRotacionView(APIView):
    """
    Endpoint para obtener el reporte de rotacion en formato JSON.

    Metodo: GET
    URL: /api/reportes/rotacion/
    Query params:
        - fecha_desde (opcional): YYYY-MM-DD
        - fecha_hasta (opcional): YYYY-MM-DD
        - insumo_id   (opcional): filtra a un solo insumo

    Respuesta exitosa (200):
    {
        "reporte": [ {...}, ... ],
        "total_insumos": int,
        "insumo_mayor_rotacion": {...} | null,
        "insumo_mayor_merma": {...} | null
    }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fecha_desde = request.query_params.get('fecha_desde') or None
        fecha_hasta = request.query_params.get('fecha_hasta') or None
        insumo_id_param = request.query_params.get('insumo_id')
        insumo_id = int(insumo_id_param) if insumo_id_param else None

        try:
            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_rotacion(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
            )

            mayor_rotacion = (
                max(reporte, key=lambda r: r['rotacion']) if reporte else None
            )
            mayor_merma = (
                max(reporte, key=lambda r: r['total_mermado']) if reporte else None
            )

            return Response({
                'reporte': reporte,
                'total_insumos': len(reporte),
                'insumo_mayor_rotacion': mayor_rotacion,
                'insumo_mayor_merma': mayor_merma,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error generando reporte de rotacion: {str(e)}")
            return Response(
                {'error': f'Error al generar el reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteRotacionPDFView(APIView):
    """
    Endpoint para descargar el reporte de rotacion en PDF.

    Metodo: GET
    URL: /api/reportes/rotacion/pdf/
    Query params:
        - fecha_desde, fecha_hasta, insumo_id (opcionales)

    Respuesta: archivo PDF (application/pdf) para descarga directa.
    Registra GENERAR_REPORTE_ROTACION en bitacora con formato='pdf'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fecha_desde = request.query_params.get('fecha_desde') or None
        fecha_hasta = request.query_params.get('fecha_hasta') or None
        insumo_id_param = request.query_params.get('insumo_id')
        insumo_id = int(insumo_id_param) if insumo_id_param else None

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_rotacion(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
            )

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=landscape(letter),
                topMargin=2 * cm, bottomMargin=2 * cm,
                leftMargin=2 * cm, rightMargin=2 * cm,
            )
            styles = getSampleStyleSheet()
            elementos = []

            titulo = "Reporte de Rotacion de Inventario"
            if fecha_desde or fecha_hasta:
                rango = f"Periodo: {fecha_desde or '...'} a {fecha_hasta or '...'}"
                titulo += f" — {rango}"

            elementos.append(Paragraph(titulo, styles['Title']))
            elementos.append(Paragraph(
                "Sistema de Gestion de Almacenes Gastronomicos · ODAA Simplificado",
                styles['Normal']
            ))
            elementos.append(Spacer(1, 0.5 * cm))

            data = [[
                'Insumo', 'Categoria',
                'Total Ingresado', 'Total Consumido', 'Total Mermado',
                'Rotacion (%)', 'Stock Actual'
            ]]
            for r in reporte:
                data.append([
                    r['nombre'],
                    r['categoria'] or '-',
                    f"{r['total_ingresado']:.2f}",
                    f"{r['total_consumido']:.2f}",
                    f"{r['total_mermado']:.2f}",
                    f"{r['rotacion']:.2f}%",
                    f"{r['stock_actual']:.2f}",
                ])

            if len(data) == 1:
                data.append(['Sin datos disponibles', '-', '-', '-', '-', '-', '-'])

            tabla = Table(
                data,
                colWidths=[5.5 * cm, 3.5 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
            )
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla)

            doc.build(elementos)
            buffer.seek(0)

            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_ROTACION",
                detalles={
                    "ip": ip_cliente,
                    "formato": "pdf",
                    "filtros": {
                        "fecha_desde": fecha_desde,
                        "fecha_hasta": fecha_hasta,
                        "insumo_id": insumo_id,
                    },
                }
            )

            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reporte_rotacion_inventario.pdf"'
            return response

        except Exception as e:
            logger.error(f"Error generando PDF de reporte de rotacion: {str(e)}")
            return Response(
                {'error': f'Error al generar el PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteRotacionExcelView(APIView):
    """
    Endpoint para descargar el reporte de rotacion en Excel.

    Metodo: GET
    URL: /api/reportes/rotacion/excel/
    Query params:
        - fecha_desde, fecha_hasta, insumo_id (opcionales)

    Respuesta: archivo .xlsx para descarga directa.
    Registra GENERAR_REPORTE_ROTACION en bitacora con formato='excel'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fecha_desde = request.query_params.get('fecha_desde') or None
        fecha_hasta = request.query_params.get('fecha_hasta') or None
        insumo_id_param = request.query_params.get('insumo_id')
        insumo_id = int(insumo_id_param) if insumo_id_param else None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_rotacion(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Rotacion"

            headers = [
                'Insumo', 'Categoria',
                'Total Ingresado', 'Total Consumido', 'Total Mermado',
                'Rotacion (%)', 'Stock Actual'
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col_num, _ in enumerate(headers, start=1):
                celda = ws.cell(row=1, column=col_num)
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = Alignment(horizontal="center")

            for r in reporte:
                ws.append([
                    r['nombre'],
                    r['categoria'] or '-',
                    r['total_ingresado'],
                    r['total_consumido'],
                    r['total_mermado'],
                    r['rotacion'],
                    r['stock_actual'],
                ])

            for col_num, header in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_num)].width = max(18, len(header) + 4)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_ROTACION",
                detalles={
                    "ip": ip_cliente,
                    "formato": "excel",
                    "filtros": {
                        "fecha_desde": fecha_desde,
                        "fecha_hasta": fecha_hasta,
                        "insumo_id": insumo_id,
                    },
                }
            )

            response = HttpResponse(
                buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="reporte_rotacion_inventario.xlsx"'
            return response

        except Exception as e:
            logger.error(f"Error generando Excel de reporte de rotacion: {str(e)}")
            return Response(
                {'error': f'Error al generar el Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
